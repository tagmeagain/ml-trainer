#!/usr/bin/env python3
"""Seed unfilled stubs from the tracker workbook (SPEC 6.1).

    python scripts/seed_from_tracker.py [--dry-run]

Reads tracker/Master_Interview_Prep_Tracker.xlsx and emits one stub per row
into content/stubs.json. A stub carries id, type, stage, category, topic and
the raw tracker text in a temporary `_source` field. It is NOT a finished card.

Card ids are permanent (CLAUDE.md rule 3), so this script is re-runnable:
existing ids are looked up by (category, topic-slug) across stubs.json AND
cards.json and reused. New rows get the next free number in their category.
Rows already promoted to cards.json are not re-stubbed.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from schema import STAGE_CATEGORY, Stub, slugify  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "tracker" / "Master_Interview_Prep_Tracker.xlsx"
CONTENT = ROOT / "content"
STUBS = CONTENT / "stubs.json"
CARDS = CONTENT / "cards.json"

# --------------------------------------------------------------------------
# Stage and category mapping (SPEC 6.1)
# --------------------------------------------------------------------------

# Tracker 2 Section column -> stage code. Category follows from the stage.
TRACKER2_SECTION = {
    "DSA — Coding Patterns": "D",
    "System Design": "SD",
    "Projects — Q&A & STAR": "B",
    "Behavioral / STAR": "B",
}

# Formulas sheet Category column -> stage code.
FORMULA_STAGE = {
    "ATTENTION": "S3", "POSITION": "S4", "NORMS": "S5", "FFN": "S5",
    "TRAINING": "S8", "PEFT": "S9", "RLHF": "S10", "QUANT": "S11",
    "GPU": "I2", "FLASHATTN": "S7", "MoE": "S12", "SOFTMAX": "S1",
    "BIAS-VAR": "S14", "A/B TEST": "S14",
}

# A few formulas belong to a different stage than their sheet category implies
# — KV cache is an inference card, not attention theory. Keyed on Name.
FORMULA_STAGE_OVERRIDES = {
    "KV Cache Memory": "I3",
    "GQA KV Reduction": "S7",
    "Spec Decoding Speedup": "I7",
    "Arithmetic Intensity": "I1",
    "LLM Decode TPOT": "I2",
    "LLM Prefill TTFT": "I2",
    "FA IO complexity": "I4",
}

SHEET_T1 = "Tracker 1 — LLM Learning"
SHEET_INF = "Inference Engineering"
SHEET_FT = "Fine-Tuning"
SHEET_T2 = "Tracker 2 — DSA+SysD+Proj"
SHEET_FORMULAS = "Formulas"

# Tracker cells use an em dash to mean "nothing to do here".
EMPTY_MARKERS = {"", "—", "-", "–", "n/a", "none"}


def _clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in EMPTY_MARKERS else text


def _stage_from_label(label: str) -> str | None:
    """'Stage 3 — Attention (the core)' -> 'S3'; 'I10 — ...' -> 'I10'."""
    text = _clean(label)
    m = re.match(r"^Stage\s+(\d+)", text, re.IGNORECASE)
    if m:
        return f"S{int(m.group(1))}"
    m = re.match(r"^([IF]\d+[A-Z]?)\b", text)
    if m:
        return m.group(1)
    return None


def _source_text(theory: str, math: str, code: str, deliverable: str = "") -> str:
    parts = []
    for label, text in (
        ("Theory", theory),
        ("Math", math),
        ("Code", code),
        ("Deliverable", deliverable),
    ):
        if text:
            parts.append(f"{label}: {text}")
    return "\n".join(parts)


# --------------------------------------------------------------------------
# Row extraction
# --------------------------------------------------------------------------


def _rows(wb, sheet: str) -> list[tuple]:
    if sheet not in wb.sheetnames:
        print(f"warning: sheet {sheet!r} not in workbook; skipped", file=sys.stderr)
        return []
    return list(wb[sheet].iter_rows(min_row=2, values_only=True))


def extract(wb) -> tuple[list[dict], Counter, list[str]]:
    """Return (raw stub dicts without ids, per-sheet counts, skip notes)."""
    out: list[dict] = []
    counts: Counter = Counter()
    notes: list[str] = []

    # -- Tracker 1: stage from the Stage column, category mapped from stage --
    # cols: # | Stage | Topic | Theory | Math | Code | Deliverable | Hrs | ...
    for row in _rows(wb, SHEET_T1):
        topic = _clean(row[2])
        if not topic:
            continue
        stage = _stage_from_label(row[1])
        if stage is None:
            notes.append(f"{SHEET_T1}: unmapped stage {row[1]!r} for topic {topic!r}")
            continue
        category = STAGE_CATEGORY.get(stage)
        if category is None:
            notes.append(f"{SHEET_T1}: no category for stage {stage}")
            continue
        out.append(
            {
                "type": "concept",
                "stage": stage,
                "category": category,
                "topic": topic,
                "_source": _source_text(_clean(row[3]), _clean(row[4]), _clean(row[5]), _clean(row[6])),
            }
        )
        counts[SHEET_T1] += 1

    # -- Inference Engineering: category `inference`, stage from the I* code --
    for row in _rows(wb, SHEET_INF):
        topic = _clean(row[2])
        if not topic:
            continue
        stage = _stage_from_label(row[1])
        if stage is None:
            notes.append(f"{SHEET_INF}: unmapped stage {row[1]!r} for topic {topic!r}")
            continue
        out.append(
            {
                "type": "concept",
                "stage": stage,
                "category": STAGE_CATEGORY[stage],
                "topic": topic,
                "_source": _source_text(_clean(row[3]), _clean(row[4]), _clean(row[5]), _clean(row[6])),
            }
        )
        counts[SHEET_INF] += 1

    # -- Fine-Tuning: category `peft`, stage from the F* code ---------------
    for row in _rows(wb, SHEET_FT):
        topic = _clean(row[2])
        if not topic:
            continue
        stage = _stage_from_label(row[1])
        if stage is None:
            notes.append(f"{SHEET_FT}: unmapped stage {row[1]!r} for topic {topic!r}")
            continue
        out.append(
            {
                "type": "concept",
                "stage": stage,
                "category": STAGE_CATEGORY[stage],
                "topic": topic,
                "_source": _source_text(_clean(row[3]), _clean(row[4]), _clean(row[5]), _clean(row[6])),
            }
        )
        counts[SHEET_FT] += 1

    # -- Tracker 2: dsa / system-design / behavioral ------------------------
    # cols: # | Section | Topic | Task | Deliverable | Hrs | Status | Notes
    for row in _rows(wb, SHEET_T2):
        topic = _clean(row[2])
        if not topic:
            continue
        section = _clean(row[1])
        stage = TRACKER2_SECTION.get(section)
        if stage is None:
            notes.append(f"{SHEET_T2}: unmapped section {section!r} for topic {topic!r}")
            continue
        category = STAGE_CATEGORY[stage]
        out.append(
            {
                "type": "concept",
                "stage": stage,
                "category": category,
                "topic": topic,
                "_source": _source_text(_clean(row[3]), "", "", _clean(row[4])),
            }
        )
        counts[SHEET_T2] += 1

    # -- Formulas: one formula card per row ---------------------------------
    # cols: Category | Name | Formula | Intuition / Key Numbers | Priority
    for row in _rows(wb, SHEET_FORMULAS):
        name = _clean(row[1])
        if not name:
            continue
        stage = FORMULA_STAGE_OVERRIDES.get(name) or FORMULA_STAGE.get(_clean(row[0]))
        if stage is None:
            notes.append(f"{SHEET_FORMULAS}: unmapped category {row[0]!r} for {name!r}")
            continue
        category = STAGE_CATEGORY[stage]
        out.append(
            {
                "type": "formula",
                "stage": stage,
                "category": category,
                "topic": name,
                "formula": _clean(row[2]),
                "priority": _clean(row[4]),
                "_source": _clean(row[3]),
            }
        )
        counts[SHEET_FORMULAS] += 1

    return out, counts, notes


# --------------------------------------------------------------------------
# Id assignment — permanent, re-runnable
# --------------------------------------------------------------------------


def _load_json(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, list) else []


def _identity(entry: dict) -> tuple[str, str, str]:
    """The stable natural key of a tracker row.

    `type` is part of it because the same topic legitimately appears twice —
    RMSNorm is both a Tracker 1 concept row and a Formulas row, and those are
    two different cards that must not collide onto one id.
    """
    cat = entry.get("category") or ""
    slug = slugify(entry.get("topic") or "")[:60].strip("-") or "topic"
    return (cat, slug, entry.get("type") or "concept")


def assign_ids(raw: list[dict]) -> tuple[list[dict], int, int]:
    """Attach permanent ids. Returns (stubs, n_reused, n_skipped_already_card)."""
    by_key: dict[tuple[str, str, str], str] = {}
    used_numbers: dict[str, set[int]] = {}
    card_keys: set[tuple[str, str, str]] = set()

    for entry in _load_json(STUBS) + _load_json(CARDS):
        cid, cat = entry.get("id"), entry.get("category")
        if not cid or not cat:
            continue
        m = re.match(rf"^{re.escape(cat)}-(.+)-(\d{{3}})$", cid)
        if not m:
            continue
        by_key[_identity(entry)] = cid
        used_numbers.setdefault(cat, set()).add(int(m.group(2)))

    for entry in _load_json(CARDS):
        card_keys.add(_identity(entry))

    stubs: list[dict] = []
    assigned: set[tuple[str, str, str]] = set()
    reused = skipped = 0

    for entry in raw:
        cat = entry["category"]
        key = _identity(entry)

        if key in card_keys:
            skipped += 1          # already a finished card, do not re-stub
            continue

        if key in by_key and key not in assigned:
            cid = by_key[key]
            reused += 1
        else:
            # New row, or a genuine within-sheet duplicate of one already
            # placed this run — either way it needs its own number.
            used = used_numbers.setdefault(cat, set())
            num = 1
            while num in used:
                num += 1
            used.add(num)
            cid = f"{cat}-{key[1]}-{num:03d}"
            by_key.setdefault(key, cid)

        assigned.add(key)
        stub = {"id": cid, **entry}
        stubs.append({k: v for k, v in stub.items() if v not in (None, "")})

    return stubs, reused, skipped


# --------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true", help="print counts, write nothing")
    args = ap.parse_args()

    if not XLSX.is_file():
        print(f"error: tracker workbook not found at {XLSX}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    try:
        raw, counts, notes = extract(wb)
    finally:
        wb.close()

    print(f"{XLSX.relative_to(ROOT)}\n")
    print(f"{'sheet':<28} {'rows':>6}")
    print("-" * 36)
    for sheet in (SHEET_T1, SHEET_INF, SHEET_FT, SHEET_T2, SHEET_FORMULAS):
        print(f"{sheet:<28} {counts[sheet]:>6}")
    print("-" * 36)
    print(f"{'TOTAL':<28} {sum(counts.values()):>6}\n")

    stubs, reused, skipped = assign_ids(raw)

    by_cat = Counter(s["category"] for s in stubs)
    print("stubs by category:")
    for cat, n in sorted(by_cat.items(), key=lambda kv: -kv[1]):
        print(f"  {cat:<20} {n:>4}")
    by_type = Counter(s["type"] for s in stubs)
    print(f"\nby type: " + ", ".join(f"{k} {v}" for k, v in sorted(by_type.items())))
    print(f"ids reused: {reused}   already finished cards (skipped): {skipped}")

    if notes:
        print(f"\n{len(notes)} unmapped row(s):", file=sys.stderr)
        for n in notes:
            print(f"  warn: {n}", file=sys.stderr)

    # Validate every stub before it touches disk.
    bad = 0
    for s in stubs:
        try:
            Stub.model_validate(s)
        except Exception as exc:
            bad += 1
            print(f"  error: stub {s.get('id')}: {exc}", file=sys.stderr)
    if bad:
        print(f"\n{bad} invalid stub(s); nothing written", file=sys.stderr)
        return 1

    if args.dry_run:
        print(f"\n--dry-run: would write {len(stubs)} stubs to {STUBS.relative_to(ROOT)}")
        return 0

    CONTENT.mkdir(exist_ok=True)
    STUBS.write_text(json.dumps(stubs, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"\nwrote {len(stubs)} stubs to {STUBS.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
